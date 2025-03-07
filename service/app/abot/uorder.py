import os
from typing import Optional
from uuid import uuid4

import aiohttp
from app.models import Payment
from app.const import UKID, UKKEY
from django.conf import settings
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


async def make_uorder(
    value, url='https://api.yookassa.ru/v3/payments', ukid=UKID, ukkey=UKKEY,
    return_url='https://www.example.com/return_url', description='Заказ №72'
) -> Optional[tuple[str, str]]:
    idmp_key = str(uuid4())
    headers = {
        'Idempotence-Key': f'{idmp_key}',
        'Content-Type': 'application/json'
    }
    data = {
        "amount": {"value": str(value), "currency": "RUB"},
        "payment_method_data": {"type": "bank_card"},
        "confirmation": {
            "type": "redirect",
            "return_url": return_url
        },
        "description": description
    }
    auth = aiohttp.BasicAuth(ukid, ukkey)
    async with aiohttp.ClientSession(auth=auth, headers=headers) as session:
        r = await session.post(url=url, json=data)
        if r.status != 200: return
        json = await r.json()
        try: conf_url = json['confirmation']['confirmation_url']
        except KeyError: return

        return conf_url, idmp_key

def export_payments_to_excel(payment: Payment, path):
    headers = ["payment_id", "final_cost", "is_paid", "created_at"]
    
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        for col_num, header in enumerate(headers, 1):
            ws[f"{get_column_letter(col_num)}1"] = header

    last_row = ws.max_row
    
    row_data = {
        "payment_id": str(payment.id),
        "final_cost": payment.final_cost,
        "is_paid": payment.is_paid,
        "created_at": str(payment.created_at)
    }
    
    last_row += 1
    for col_num, header in enumerate(headers, 1):
        ws[f"{get_column_letter(col_num)}{last_row}"] = row_data[header]

    wb.save(path)

async def after_payment(payment_uid):
    path = f'{settings.BASE_DIR}/payments.xlsx'
    if not (payment:=await Payment.objects.filter(pk=payment_uid).afirst()): return
    payment.is_paid = True
    await payment.asave()
    export_payments_to_excel(payment, path)
